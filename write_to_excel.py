import openpyxl
import pandas as pd

exPath = "C:\\Users\\aksha\\Desktop\\Virtusa\\test.xlsx" #excel file path
pyPath = "C:\\Users\\aksha\\Desktop\\Virtusa\\dag.py"	 #dag file path
file = open(pyPath)
search = ['hdfs_path=', 'task_name=']

bundle = ''
hdfs = ''
x = 0
j = 1
c = 0
num_lines = 0

with open("C:\\Users\\aksha\\Desktop\\Virtusa\\dag.py", 'r') as fp:
    x = len(fp.readlines())
fp.close()

while(j <= x):
	s = file.readline()
	list = s.split()
	if (search[0] in list):
		hdfs = list[1]
		c = c + 1

	if (search[1] in list):
		bundle = list[1]
		c = c + 1

	if (c == len(search)):
		break

	j = j + 1

bundle = bundle[1:len(bundle)-1]
scriptName = bundle + '.hql'
hdfsPath = hdfs[1:len(hdfs)-1]
list = hdfsPath.split('/')
hdfsPath = hdfsPath + '/' + scriptName
coord = list[2]

print("Script Name: ", scriptName)
print("HDFS Path of Script: ", hdfsPath)

wb = openpyxl.load_workbook(exPath)
ws = wb['Sheet1']

df = pd.read_excel(exPath)
i = df.shape[0]

def writeToExcel(i, bundle, coord, hdfsPath, scriptName):
	ws.cell(row=i,column=1,value=bundle)
	ws.cell(row=i,column=2,value=coord)
	ws.cell(row=i,column=6,value='hive')
	ws.cell(row=i,column=7,value=hdfsPath)
	ws.cell(row=i,column=8,value=scriptName)

writeToExcel(i+2, bundle, coord, hdfsPath, scriptName)

wb.save(exPath)